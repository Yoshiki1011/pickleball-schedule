#!/usr/bin/env python3
# ============================================================
# ピックルボールダブルス 紅白対抗戦 Excel出力ツール（汎用版）
# ============================================================
# 使い方:
#   python3 make_excel_universal.py
#   （schedule_data.json を読み込んで Excel を生成します）
#
# 事前に pickleball_generator.js を実行して
# schedule_data.json を生成してください。
# ============================================================

import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ===== JSONデータ読み込み =====
json_path = '/Users/yoshiki.nagatome/Claude/Doc/schedule_data.json'
with open(json_path, encoding='utf-8') as f:
    data = json.load(f)

schedule   = data['schedule']
params     = data['params']
HALF       = params['half']
COURTS     = params['courts']
TOTAL_GAMES = params['totalGames']
GAMES_PER  = params['gamesPerPlayer']
REST_PER   = params['restPerPlayer']

print(f"読み込み完了: {HALF*2}人・{COURTS}面・{TOTAL_GAMES}試合")

# ===== カラー定義 =====
RED_LIGHT     = PatternFill("solid", fgColor="FFCCCC")
WHITE_LIGHT   = PatternFill("solid", fgColor="DDDDFF")
REST_RED_FILL = PatternFill("solid", fgColor="FF9999")
REST_WHT_FILL = PatternFill("solid", fgColor="9999FF")
HEADER_FILL   = PatternFill("solid", fgColor="2F2F2F")
COURT_FILL    = PatternFill("solid", fgColor="FFF2CC")
TITLE_FILL    = PatternFill("solid", fgColor="1F4E79")
SUB_FILL      = PatternFill("solid", fgColor="2E75B6")
GAME_FILL     = PatternFill("solid", fgColor="1F4E79")
VS_FILL       = PatternFill("solid", fgColor="F2F2F2")
SEP_FILL      = PatternFill("solid", fgColor="CCCCCC")
LEGEND_FILL   = PatternFill("solid", fgColor="F5F5F5")

THIN = Side(style="thin",   color="AAAAAA")
MED  = Side(style="medium", color="555555")

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ===== 列構成（コート数に応じて動的生成）=====
# 固定列: 試合番号(1), コート(2)
# コートごと: 紅①(3), 紅②(4), vs(5), 白①(6), 白②(7) → 5列 × COURTS
# 区切り(1列)
# 紅休み × rest_per_side 列
# 区切り(1列)
# 白休み × rest_per_side 列

rest_per_side = (HALF - COURTS * 2) // 2  # 片組の休み人数
# 実際はJSONから取得
rest_per_side = len(schedule[0]['redRest'])

FIXED_COLS = 2
COURT_COLS = 5  # 紅①・紅②・vs・白①・白②
SEP_COL    = 1
REST_COLS  = rest_per_side

total_cols = FIXED_COLS + COURT_COLS * COURTS + SEP_COL + REST_COLS + SEP_COL + REST_COLS
last_col_letter = get_column_letter(total_cols)

# ===== ワークブック作成 =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "紅白対抗戦スケジュール"

# ===== タイトル =====
ws.merge_cells(f"A1:{last_col_letter}1")
c = ws.cell(row=1, column=1,
    value=f"🏓  ピックルボールダブルス  紅白対抗戦スケジュール")
c.fill = TITLE_FILL
c.font = Font(bold=True, size=16, color="FFFFFF")
c.alignment = CENTER
ws.row_dimensions[1].height = 38

ws.merge_cells(f"A2:{last_col_letter}2")
c = ws.cell(row=2, column=1,
    value=f"{HALF*2}人（紅組1〜{HALF}番・白組{HALF+1}〜{HALF*2}番）｜ {COURTS}面 ｜ {TOTAL_GAMES}試合 ｜ 全員{GAMES_PER}試合出場・{REST_PER}試合休み ｜ 同じ人と組むのは最大2回まで ｜ 対戦相手も分散")
c.fill = SUB_FILL
c.font = Font(bold=True, size=10, color="FFFFFF")
c.alignment = CENTER
ws.row_dimensions[2].height = 20

# ===== ヘッダー行 =====
row = 3
col = 1

# 試合
c = ws.cell(row=row, column=col, value="試合"); c.fill = HEADER_FILL; c.font = Font(bold=True, color="FFFFFF", size=9); c.alignment = CENTER; c.border = thin_border()
col += 1

# コート
c = ws.cell(row=row, column=col, value="コート"); c.fill = HEADER_FILL; c.font = Font(bold=True, color="FFFFFF", size=9); c.alignment = CENTER; c.border = thin_border()
col += 1

# コートごとのヘッダー
for ct in range(COURTS):
    for label in [f"紅①", f"紅②", "vs", f"白①", f"白②"]:
        c = ws.cell(row=row, column=col, value=label)
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = CENTER
        c.border = thin_border()
        col += 1

# 区切り
c = ws.cell(row=row, column=col, value=""); c.fill = HEADER_FILL; c.border = thin_border(); col += 1

# 紅休み
for i in range(rest_per_side):
    c = ws.cell(row=row, column=col, value=f"紅\n休み{i+1}")
    c.fill = HEADER_FILL; c.font = Font(bold=True, color="FFFFFF", size=9); c.alignment = CENTER; c.border = thin_border()
    col += 1

# 区切り
c = ws.cell(row=row, column=col, value=""); c.fill = HEADER_FILL; c.border = thin_border(); col += 1

# 白休み
for i in range(rest_per_side):
    c = ws.cell(row=row, column=col, value=f"白\n休み{i+1}")
    c.fill = HEADER_FILL; c.font = Font(bold=True, color="FFFFFF", size=9); c.alignment = CENTER; c.border = thin_border()
    col += 1

ws.row_dimensions[row].height = 28

# ===== データ行 =====
current_row = 4
court_labels = [f"コート{i+1}" for i in range(COURTS)]

for game_data in schedule:
    game_num   = game_data['game']
    red_rest   = game_data['redRest']
    white_rest = game_data['whiteRest']
    courts     = game_data['courts']

    for i in range(COURTS):
        r = current_row + i
        is_first = (i == 0)

        col = 1

        # 試合番号（マージ予定）
        c = ws.cell(row=r, column=col, value=f"試合\n{game_num}" if is_first else "")
        c.fill = GAME_FILL; c.font = Font(bold=True, color="FFFFFF", size=11); c.alignment = CENTER; c.border = thin_border()
        col += 1

        # コート
        c = ws.cell(row=r, column=col, value=court_labels[i])
        c.fill = COURT_FILL; c.font = Font(bold=True, size=9, color="555500"); c.alignment = CENTER; c.border = thin_border()
        col += 1

        # コートごとのペア
        for ct in range(COURTS):
            if ct < len(courts):
                rp = courts[ct]['red']
                wp = courts[ct]['white']
            else:
                rp = ['-', '-']
                wp = ['-', '-']

            if ct == i:
                # このコートの試合
                c = ws.cell(row=r, column=col, value=f"紅{rp[0]}"); c.fill = RED_LIGHT; c.font = Font(color="CC0000", bold=True, size=12); c.alignment = CENTER; c.border = thin_border(); col += 1
                c = ws.cell(row=r, column=col, value=f"紅{rp[1]}"); c.fill = RED_LIGHT; c.font = Font(color="CC0000", bold=True, size=12); c.alignment = CENTER; c.border = thin_border(); col += 1
                c = ws.cell(row=r, column=col, value="vs"); c.fill = VS_FILL; c.font = Font(bold=True, size=9, color="888888"); c.alignment = CENTER; c.border = thin_border(); col += 1
                c = ws.cell(row=r, column=col, value=f"白{wp[0]}"); c.fill = WHITE_LIGHT; c.font = Font(color="0000CC", bold=True, size=12); c.alignment = CENTER; c.border = thin_border(); col += 1
                c = ws.cell(row=r, column=col, value=f"白{wp[1]}"); c.fill = WHITE_LIGHT; c.font = Font(color="0000CC", bold=True, size=12); c.alignment = CENTER; c.border = thin_border(); col += 1
            else:
                for _ in range(5):
                    c = ws.cell(row=r, column=col, value=""); c.fill = PatternFill("solid", fgColor="F8F8F8"); c.border = thin_border(); col += 1

        # 区切り
        c = ws.cell(row=r, column=col, value=""); c.fill = SEP_FILL; c.border = thin_border(); col += 1

        # 紅休み（マージ予定）
        for j in range(rest_per_side):
            val = f"紅{red_rest[j]}" if is_first else ""
            c = ws.cell(row=r, column=col, value=val)
            c.fill = REST_RED_FILL; c.font = Font(color="880000", bold=True, size=11); c.alignment = CENTER; c.border = thin_border()
            col += 1

        # 区切り
        c = ws.cell(row=r, column=col, value=""); c.fill = SEP_FILL; c.border = thin_border(); col += 1

        # 白休み（マージ予定）
        for j in range(rest_per_side):
            val = f"白{white_rest[j]}" if is_first else ""
            c = ws.cell(row=r, column=col, value=val)
            c.fill = REST_WHT_FILL; c.font = Font(color="000088", bold=True, size=11); c.alignment = CENTER; c.border = thin_border()
            col += 1

        ws.row_dimensions[r].height = 24

    # マージ（試合番号・休み欄）
    merge_cols = [1]  # 試合番号
    # 休み列のマージ
    rest_start_col = FIXED_COLS + COURT_COLS * COURTS + SEP_COL + 1
    for j in range(rest_per_side):
        merge_cols.append(rest_start_col + j)
    rest_start_col2 = rest_start_col + rest_per_side + SEP_COL
    for j in range(rest_per_side):
        merge_cols.append(rest_start_col2 + j)

    for mc in merge_cols:
        ws.merge_cells(start_row=current_row, start_column=mc, end_row=current_row + COURTS - 1, end_column=mc)

    # 試合間の区切り線
    for c_idx in range(1, total_cols + 1):
        cell = ws.cell(row=current_row + COURTS - 1, column=c_idx)
        existing = cell.border
        cell.border = Border(left=existing.left, right=existing.right, top=existing.top, bottom=MED)

    current_row += COURTS

# ===== 列幅設定 =====
col = 1
ws.column_dimensions[get_column_letter(col)].width = 7; col += 1   # 試合
ws.column_dimensions[get_column_letter(col)].width = 9; col += 1   # コート
for ct in range(COURTS):
    ws.column_dimensions[get_column_letter(col)].width = 7; col += 1   # 紅①
    ws.column_dimensions[get_column_letter(col)].width = 7; col += 1   # 紅②
    ws.column_dimensions[get_column_letter(col)].width = 4; col += 1   # vs
    ws.column_dimensions[get_column_letter(col)].width = 7; col += 1   # 白①
    ws.column_dimensions[get_column_letter(col)].width = 7; col += 1   # 白②
ws.column_dimensions[get_column_letter(col)].width = 2; col += 1   # 区切り
for j in range(rest_per_side):
    ws.column_dimensions[get_column_letter(col)].width = 7; col += 1
ws.column_dimensions[get_column_letter(col)].width = 2; col += 1   # 区切り
for j in range(rest_per_side):
    ws.column_dimensions[get_column_letter(col)].width = 7; col += 1

# ===== 凡例 =====
legend_row = current_row + 1
ws.merge_cells(f"A{legend_row}:{last_col_letter}{legend_row}")
c = ws.cell(row=legend_row, column=1,
    value="【凡例】 紅①・紅② = 紅組ペア番号（赤背景）　白①・白② = 白組ペア番号（青背景）　紅休み・白休み = その試合を休む選手番号")
c.fill = LEGEND_FILL; c.font = Font(size=9, color="444444", italic=True); c.alignment = LEFT
ws.row_dimensions[legend_row].height = 16

# ===== シート設定 =====
ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False

# ===== 保存 =====
output_path = '/Users/yoshiki.nagatome/Claude/Doc/ピックルボール紅白対抗戦スケジュール.xlsx'
wb.save(output_path)
print(f"✅ 保存完了: {output_path}")
