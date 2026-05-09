import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "紅白対抗戦スケジュール"

# ===== カラー定義 =====
RED_FILL      = PatternFill("solid", fgColor="FF6666")   # 紅組
RED_LIGHT     = PatternFill("solid", fgColor="FFCCCC")   # 紅組（薄）
WHITE_FILL    = PatternFill("solid", fgColor="AAAAFF")   # 白組
WHITE_LIGHT   = PatternFill("solid", fgColor="DDDDFF")   # 白組（薄）
REST_FILL     = PatternFill("solid", fgColor="DDDDDD")   # 休み
HEADER_FILL   = PatternFill("solid", fgColor="333333")   # ヘッダー背景
COURT_FILL    = PatternFill("solid", fgColor="FFF2CC")   # コート行
TITLE_FILL    = PatternFill("solid", fgColor="1F4E79")   # タイトル

THIN = Side(style="thin", color="999999")
THICK = Side(style="medium", color="444444")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

def cell_style(ws, row, col, value, fill=None, font=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill = fill
    if font:   c.font = font
    if align:  c.alignment = align
    if border: c.border = border
    return c

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ===== スケジュールデータ =====
schedule = [
    # (試合番号, 紅休み, 白休み, [(紅ペア, 白ペア), ...])
    (1,  [1,2],   [9,10],  [([3,4],[11,12]), ([5,6],[13,14]), ([7,8],[15,16])]),
    (2,  [3,4],   [11,12], [([1,2],[9,10]),  ([5,7],[13,15]), ([6,8],[14,16])]),
    (3,  [5,6],   [13,14], [([1,3],[9,11]),  ([2,4],[10,12]), ([7,8],[15,16])]),
    (4,  [7,8],   [15,16], [([1,4],[9,12]),  ([2,3],[10,11]), ([5,6],[13,14])]),
    (5,  [1,2],   [9,10],  [([3,5],[11,13]), ([4,7],[12,15]), ([6,8],[14,16])]),
    (6,  [3,4],   [11,12], [([1,5],[9,13]),  ([2,8],[10,16]), ([6,7],[14,15])]),
    (7,  [5,6],   [13,14], [([1,7],[9,15]),  ([2,3],[10,11]), ([4,8],[12,16])]),
    (8,  [7,8],   [15,16], [([1,6],[9,14]),  ([2,5],[10,13]), ([3,4],[11,12])]),
    (9,  [1,2],   [9,10],  [([3,6],[11,14]), ([4,7],[12,15]), ([5,8],[13,16])]),
    (10, [3,4],   [11,12], [([1,8],[9,16]),  ([2,6],[10,14]), ([5,7],[13,15])]),
    (11, [5,6],   [13,14], [([1,2],[9,10]),  ([3,7],[11,15]), ([4,8],[12,16])]),
    (12, [7,8],   [15,16], [([1,3],[9,11]),  ([2,5],[10,13]), ([4,6],[12,14])]),
]

# ===== タイトル行 =====
ws.merge_cells("A1:M1")
c = ws.cell(row=1, column=1, value="🏓 ピックルボールダブルス 紅白対抗戦スケジュール")
c.fill = TITLE_FILL
c.font = Font(bold=True, size=16, color="FFFFFF")
c.alignment = CENTER
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:M2")
c = ws.cell(row=2, column=1, value="16人・3面・12試合　｜　全員9試合出場・3試合休み　｜　同じ人と組むのは最大2回まで")
c.fill = PatternFill("solid", fgColor="2E75B6")
c.font = Font(bold=True, size=11, color="FFFFFF")
c.alignment = CENTER
ws.row_dimensions[2].height = 22

# ===== ヘッダー行 =====
# 列構成:
# A: 試合番号
# B: コート
# C-D: 紅ペア
# E: vs
# F-G: 白ペア
# H: (空白 or 区切り)
# I-J: 紅休み
# K: /
# L-M: 白休み

headers = [
    (1, "試合"),
    (2, "コート"),
    (3, "紅①"),
    (4, "紅②"),
    (5, "vs"),
    (6, "白①"),
    (7, "白②"),
    (8, ""),
    (9, "紅休み①"),
    (10, "紅休み②"),
    (11, ""),
    (12, "白休み①"),
    (13, "白休み②"),
]

row = 3
for col, label in headers:
    c = ws.cell(row=row, column=col, value=label)
    c.fill = HEADER_FILL
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.alignment = CENTER
    c.border = THIN_BORDER
ws.row_dimensions[row].height = 22

# ===== データ行 =====
current_row = 4
game_start_rows = []

for game_num, red_rest, white_rest, courts in schedule:
    game_start_rows.append(current_row)
    court_labels = ["コート1", "コート2", "コート3"]

    for i, (red_pair, white_pair) in enumerate(courts):
        r = current_row + i

        # 試合番号（3行マージ予定）
        c = ws.cell(row=r, column=1, value=f"試合{game_num}" if i == 0 else "")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = CENTER
        c.border = THIN_BORDER

        # コート
        c = ws.cell(row=r, column=2, value=court_labels[i])
        c.fill = COURT_FILL
        c.font = Font(bold=True, size=10)
        c.alignment = CENTER
        c.border = THIN_BORDER

        # 紅ペア
        c = ws.cell(row=r, column=3, value=f"紅{red_pair[0]}")
        c.fill = RED_LIGHT
        c.font = Font(color="CC0000", bold=True, size=11)
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = ws.cell(row=r, column=4, value=f"紅{red_pair[1]}")
        c.fill = RED_LIGHT
        c.font = Font(color="CC0000", bold=True, size=11)
        c.alignment = CENTER
        c.border = THIN_BORDER

        # vs
        c = ws.cell(row=r, column=5, value="vs")
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.font = Font(bold=True, size=10, color="666666")
        c.alignment = CENTER
        c.border = THIN_BORDER

        # 白ペア
        c = ws.cell(row=r, column=6, value=f"白{white_pair[0]}")
        c.fill = WHITE_LIGHT
        c.font = Font(color="0000CC", bold=True, size=11)
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = ws.cell(row=r, column=7, value=f"白{white_pair[1]}")
        c.fill = WHITE_LIGHT
        c.font = Font(color="0000CC", bold=True, size=11)
        c.alignment = CENTER
        c.border = THIN_BORDER

        # 区切り
        c = ws.cell(row=r, column=8, value="")
        c.fill = PatternFill("solid", fgColor="EEEEEE")
        c.border = THIN_BORDER

        # 休み（3行目のみ表示、マージ予定）
        c = ws.cell(row=r, column=9, value=f"紅{red_rest[0]}" if i == 0 else "")
        c.fill = REST_FILL
        c.font = Font(color="CC0000", size=10)
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = ws.cell(row=r, column=10, value=f"紅{red_rest[1]}" if i == 0 else "")
        c.fill = REST_FILL
        c.font = Font(color="CC0000", size=10)
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = ws.cell(row=r, column=11, value="休" if i == 0 else "")
        c.fill = REST_FILL
        c.font = Font(bold=True, size=9, color="666666")
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = ws.cell(row=r, column=12, value=f"白{white_rest[0]}" if i == 0 else "")
        c.fill = REST_FILL
        c.font = Font(color="0000CC", size=10)
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = ws.cell(row=r, column=13, value=f"白{white_rest[1]}" if i == 0 else "")
        c.fill = REST_FILL
        c.font = Font(color="0000CC", size=10)
        c.alignment = CENTER
        c.border = THIN_BORDER

        ws.row_dimensions[r].height = 22

    # 試合番号セルをマージ
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=1)
    ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row+2, end_column=9)
    ws.merge_cells(start_row=current_row, start_column=10, end_row=current_row+2, end_column=10)
    ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row+2, end_column=11)
    ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row+2, end_column=12)
    ws.merge_cells(start_row=current_row, start_column=13, end_row=current_row+2, end_column=13)

    # 試合間の区切り線（太め）
    for col in range(1, 14):
        c = ws.cell(row=current_row + 2, column=col)
        c.border = Border(
            left=c.border.left,
            right=c.border.right,
            top=c.border.top,
            bottom=Side(style="medium", color="888888")
        )

    current_row += 3

# ===== 列幅設定 =====
col_widths = {
    1: 9,   # 試合
    2: 9,   # コート
    3: 7,   # 紅①
    4: 7,   # 紅②
    5: 5,   # vs
    6: 7,   # 白①
    7: 7,   # 白②
    8: 3,   # 区切り
    9: 8,   # 紅休み①
    10: 8,  # 紅休み②
    11: 4,  # 休
    12: 8,  # 白休み①
    13: 8,  # 白休み②
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ===== 凡例 =====
legend_row = current_row + 1
ws.merge_cells(f"A{legend_row}:M{legend_row}")
c = ws.cell(row=legend_row, column=1,
            value="【凡例】 紅① 紅② = 紅組ペア（番号）　白① 白② = 白組ペア（番号）　休み欄 = その試合を休む選手番号")
c.fill = PatternFill("solid", fgColor="F2F2F2")
c.font = Font(size=9, color="444444", italic=True)
c.alignment = LEFT
ws.row_dimensions[legend_row].height = 18

# ===== シート設定 =====
ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False

# ===== 保存 =====
output_path = "/Users/yoshiki.nagatome/Claude/Doc/ピックルボール紅白対抗戦スケジュール.xlsx"
wb.save(output_path)
print(f"保存完了: {output_path}")
