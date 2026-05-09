import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# JSONデータ読み込み
with open('/Users/yoshiki.nagatome/Claude/Doc/schedule_data.json', encoding='utf-8') as f:
    data = json.load(f)

schedule = data['schedule']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "紅白対抗戦スケジュール"

# ===== カラー定義 =====
RED_LIGHT     = PatternFill("solid", fgColor="FFCCCC")
WHITE_LIGHT   = PatternFill("solid", fgColor="DDDDFF")
REST_RED_FILL = PatternFill("solid", fgColor="FF9999")
REST_WHT_FILL = PatternFill("solid", fgColor="9999FF")
HEADER_FILL   = PatternFill("solid", fgColor="2F2F2F")
COURT_FILL    = PatternFill("solid", fgColor="FFF2CC")
TITLE_FILL    = PatternFill("solid", fgColor="1F4E79")
SUB_FILL      = PatternFill("solid", fgColor="2E75B6")
GAME_FILL     = PatternFill("solid", fgColor="1F4E79")
VS_FILL       = PatternFill("solid", fgColor="F2F2F2")
SEP_FILL      = PatternFill("solid", fgColor="CCCCCC")
LEGEND_FILL   = PatternFill("solid", fgColor="F5F5F5")

THIN  = Side(style="thin",   color="AAAAAA")
MED   = Side(style="medium", color="666666")
THICK = Side(style="medium", color="333333")

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def bottom_thick():
    return Border(left=THIN, right=THIN, top=THIN, bottom=MED)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ===== 列構成 =====
# A(1): 試合番号
# B(2): コート
# C(3): 紅①
# D(4): 紅②
# E(5): vs
# F(6): 白①
# G(7): 白②
# H(8): 区切り
# I(9): 紅休み①
# J(10): 紅休み②
# K(11): 区切り
# L(12): 白休み①
# M(13): 白休み②

# ===== タイトル =====
ws.merge_cells("A1:M1")
c = ws.cell(row=1, column=1, value="🏓  ピックルボールダブルス  紅白対抗戦スケジュール")
c.fill = TITLE_FILL
c.font = Font(bold=True, size=16, color="FFFFFF")
c.alignment = CENTER
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:M2")
c = ws.cell(row=2, column=1,
    value="16人（紅組1〜8番・白組9〜16番）｜ 3面 ｜ 12試合 ｜ 全員9試合出場・3試合休み ｜ 同じ人と組むのは最大2回まで ｜ 対戦相手も分散")
c.fill = SUB_FILL
c.font = Font(bold=True, size=10, color="FFFFFF")
c.alignment = CENTER
ws.row_dimensions[2].height = 20

# ===== ヘッダー =====
headers = [
    (1, "試合"),
    (2, "コート"),
    (3, "紅①"),
    (4, "紅②"),
    (5, "vs"),
    (6, "白①"),
    (7, "白②"),
    (8, ""),
    (9, "紅\n休み①"),
    (10, "紅\n休み②"),
    (11, ""),
    (12, "白\n休み①"),
    (13, "白\n休み②"),
]
row = 3
for col, label in headers:
    c = ws.cell(row=row, column=col, value=label)
    c.fill = HEADER_FILL
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.alignment = CENTER
    c.border = thin_border()
ws.row_dimensions[row].height = 28

# ===== データ行 =====
current_row = 4
court_labels = ["コート1", "コート2", "コート3"]

for game_data in schedule:
    game_num   = game_data['game']
    red_rest   = game_data['redRest']
    white_rest = game_data['whiteRest']
    courts     = game_data['courts']

    for i, court in enumerate(courts):
        r = current_row + i
        rp = court['red']
        wp = court['white']
        is_first = (i == 0)
        is_last  = (i == 2)

        # 試合番号（マージ予定）
        c = ws.cell(row=r, column=1, value=f"試合\n{game_num}" if is_first else "")
        c.fill = GAME_FILL
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = CENTER
        c.border = thin_border()

        # コート
        c = ws.cell(row=r, column=2, value=court_labels[i])
        c.fill = COURT_FILL
        c.font = Font(bold=True, size=9, color="555500")
        c.alignment = CENTER
        c.border = thin_border()

        # 紅ペア
        for col_idx, player in zip([3, 4], rp):
            c = ws.cell(row=r, column=col_idx, value=f"紅{player}")
            c.fill = RED_LIGHT
            c.font = Font(color="CC0000", bold=True, size=12)
            c.alignment = CENTER
            c.border = thin_border()

        # vs
        c = ws.cell(row=r, column=5, value="vs")
        c.fill = VS_FILL
        c.font = Font(bold=True, size=9, color="888888")
        c.alignment = CENTER
        c.border = thin_border()

        # 白ペア
        for col_idx, player in zip([6, 7], wp):
            c = ws.cell(row=r, column=col_idx, value=f"白{player}")
            c.fill = WHITE_LIGHT
            c.font = Font(color="0000CC", bold=True, size=12)
            c.alignment = CENTER
            c.border = thin_border()

        # 区切り
        c = ws.cell(row=r, column=8, value="")
        c.fill = SEP_FILL
        c.border = thin_border()

        # 紅休み（マージ予定）
        c = ws.cell(row=r, column=9, value=f"紅{red_rest[0]}" if is_first else "")
        c.fill = REST_RED_FILL
        c.font = Font(color="880000", bold=True, size=11)
        c.alignment = CENTER
        c.border = thin_border()

        c = ws.cell(row=r, column=10, value=f"紅{red_rest[1]}" if is_first else "")
        c.fill = REST_RED_FILL
        c.font = Font(color="880000", bold=True, size=11)
        c.alignment = CENTER
        c.border = thin_border()

        # 区切り
        c = ws.cell(row=r, column=11, value="")
        c.fill = SEP_FILL
        c.border = thin_border()

        # 白休み（マージ予定）
        c = ws.cell(row=r, column=12, value=f"白{white_rest[0]}" if is_first else "")
        c.fill = REST_WHT_FILL
        c.font = Font(color="000088", bold=True, size=11)
        c.alignment = CENTER
        c.border = thin_border()

        c = ws.cell(row=r, column=13, value=f"白{white_rest[1]}" if is_first else "")
        c.fill = REST_WHT_FILL
        c.font = Font(color="000088", bold=True, size=11)
        c.alignment = CENTER
        c.border = thin_border()

        ws.row_dimensions[r].height = 24

    # マージ
    ws.merge_cells(start_row=current_row, start_column=1,  end_row=current_row+2, end_column=1)
    ws.merge_cells(start_row=current_row, start_column=9,  end_row=current_row+2, end_column=9)
    ws.merge_cells(start_row=current_row, start_column=10, end_row=current_row+2, end_column=10)
    ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row+2, end_column=11)
    ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row+2, end_column=12)
    ws.merge_cells(start_row=current_row, start_column=13, end_row=current_row+2, end_column=13)

    # 試合間の区切り線（太め）
    for col in range(1, 14):
        c = ws.cell(row=current_row + 2, column=col)
        existing = c.border
        c.border = Border(
            left=existing.left,
            right=existing.right,
            top=existing.top,
            bottom=Side(style="medium", color="555555")
        )

    current_row += 3

# ===== 列幅 =====
col_widths = {
    1: 7,    # 試合
    2: 9,    # コート
    3: 7,    # 紅①
    4: 7,    # 紅②
    5: 4.5,  # vs
    6: 7,    # 白①
    7: 7,    # 白②
    8: 2,    # 区切り
    9: 7,    # 紅休み①
    10: 7,   # 紅休み②
    11: 2,   # 区切り
    12: 7,   # 白休み①
    13: 7,   # 白休み②
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ===== 凡例 =====
legend_row = current_row + 1
ws.merge_cells(f"A{legend_row}:M{legend_row}")
c = ws.cell(row=legend_row, column=1,
    value="【凡例】 紅①・紅② = 紅組ペア番号（赤背景）　白①・白② = 白組ペア番号（青背景）　紅休み・白休み = その試合を休む選手番号")
c.fill = LEGEND_FILL
c.font = Font(size=9, color="444444", italic=True)
c.alignment = LEFT
ws.row_dimensions[legend_row].height = 16

# ===== シート設定 =====
ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False

# ===== 保存 =====
output_path = "/Users/yoshiki.nagatome/Claude/Doc/ピックルボール紅白対抗戦スケジュール.xlsx"
wb.save(output_path)
print(f"保存完了: {output_path}")
